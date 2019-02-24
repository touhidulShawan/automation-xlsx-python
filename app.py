import openpyxl as xl
from openpyxl.chart import BarChart, Reference


def processing_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        discount_price = cell.value * 0.3
        discount_price_sale = sheet.cell(row, 4)
        discount_price_sale.value = discount_price

    values = Reference(
        sheet,
        min_row=2,
        max_row=sheet.max_row,
        min_col=4,
        max_col=4
    )

    bar_chart = BarChart()
    bar_chart.add_data(values)
    sheet.add_chart(bar_chart, 'E2')

    wb.save(filename)


processing_workbook('Sales_record.xlsx')
